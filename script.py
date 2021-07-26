import openpyxl

# pip install -r requirements.txt
# source .venv/bin/activate

number_of_rows = 61
wb = openpyxl.load_workbook('./localize.xlsx')
sheet = wb["list-1"]

def parse_data():
    data = ""
    for row in range(2, number_of_rows + 1):
        ru = sheet.cell(row=row, column=1).value
        ua = sheet.cell(row=row, column=2).value
        str = '"{}" = "{}";\n'.format(ru, ua)
        data += str
    save_data(data)

def save_data(str):
    with open('datat.txt', 'w') as file:
        file.write(str)
        wb.close()

parse_data()